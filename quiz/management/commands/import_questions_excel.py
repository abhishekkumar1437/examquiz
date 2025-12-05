import os
import re
from pathlib import Path
from django.core.management.base import BaseCommand
from django.conf import settings
from django.db import transaction
from quiz.models import Category, Exam, Topic, Question, Choice

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class Command(BaseCommand):
    help = 'Import questions from Excel files (.xlsx, .xls) in the inbox folder'

    def add_arguments(self, parser):
        parser.add_argument(
            '--inbox-folder',
            type=str,
            default='inbox',
            help='Folder path relative to project root containing Excel files (default: inbox)'
        )
        parser.add_argument(
            '--sheet-name',
            type=str,
            default=None,
            help='Specific sheet name to import (default: first sheet)'
        )
        parser.add_argument(
            '--header-row',
            type=int,
            default=1,
            help='Row number containing headers (1-based, default: 1)'
        )
        parser.add_argument(
            '--start-row',
            type=int,
            default=2,
            help='Row number where data starts (1-based, default: 2)'
        )

    def handle(self, *args, **options):
        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            self.stdout.write(self.style.ERROR(
                'Error: Neither openpyxl nor pandas is installed. '
                'Please install one of them: pip install openpyxl or pip install pandas'
            ))
            return

        inbox_folder = options['inbox_folder']
        sheet_name = options['sheet_name']
        header_row = options['header_row']
        start_row = options['start_row']
        
        # Get absolute path to inbox folder
        base_dir = Path(settings.BASE_DIR)
        inbox_path = base_dir / inbox_folder
        
        # Create inbox, processed, and failed folders if they don't exist
        inbox_path.mkdir(exist_ok=True)
        processed_path = inbox_path.parent / 'processed'
        failed_path = inbox_path.parent / 'failed'
        processed_path.mkdir(exist_ok=True)
        failed_path.mkdir(exist_ok=True)
        
        self.stdout.write(self.style.SUCCESS(f'Looking for Excel files in: {inbox_path}'))
        
        # Find all Excel files
        excel_files = list(inbox_path.glob('*.xlsx')) + list(inbox_path.glob('*.xls'))
        
        if not excel_files:
            self.stdout.write(self.style.WARNING('No Excel files (.xlsx or .xls) found in inbox folder.'))
            return
        
        self.stdout.write(self.style.SUCCESS(f'Found {len(excel_files)} Excel file(s) to process.'))
        
        # Process each Excel file
        for excel_file in excel_files:
            self.stdout.write(f'\nProcessing: {excel_file.name}')
            success = self.process_excel_file(excel_file, processed_path, failed_path, sheet_name, header_row, start_row)
            
            if success:
                self.stdout.write(self.style.SUCCESS(f'Successfully processed {excel_file.name}'))
            else:
                self.stdout.write(self.style.ERROR(f'Failed to process {excel_file.name}'))

    def normalize_column_name(self, col_name):
        """Normalize column name to lowercase and remove spaces"""
        if not col_name:
            return None
        # Convert to lowercase and strip whitespace
        normalized = str(col_name).lower().strip()
        # Replace spaces with underscores
        normalized = normalized.replace(' ', '_')
        # Remove brackets if present
        normalized = normalized.replace('[', '').replace(']', '')
        return normalized

    def map_excel_columns(self, fieldnames):
        """Map Excel column names to standard format (case-insensitive)"""
        column_map = {}
        normalized_fieldnames = {self.normalize_column_name(col): col for col in fieldnames}
        
        # Direct mapping with variations
        column_mappings = {
            'exam_category': ['exam_category', 'examcategory', 'exam_cat'],
            'category': ['category', 'cat'],
            'exam_name': ['exam_name', 'examname'],
            'exam': ['exam', 'exam_set'],
            'question_text': ['question_text', 'question', 'questiontext', 'q'],
            'topic': ['topic'],
            'difficulty': ['difficulty', 'diff'],
            'explanation': ['explanation', 'expl'],
            'points': ['points', 'point'],
            'question_type': ['question_type', 'type', 'qtype'],
            'correct_answer': ['correct_answer', 'correct_choices', 'correct', 'answer', 'right_answer'],
        }
        
        # Map standard columns
        for standard_name, variations in column_mappings.items():
            for variation in variations:
                if variation in normalized_fieldnames:
                    column_map[standard_name] = normalized_fieldnames[variation]
                    break
        
        # Map choice columns (Choice 1, Choice 2, etc.)
        for i in range(1, 7):
            choice_variations = [f'choice_{i}', f'choice{i}', f'option_{i}', f'option{i}']
            for variation in choice_variations:
                if variation in normalized_fieldnames:
                    column_map[f'choice_{i}'] = normalized_fieldnames[variation]
                    break
        
        # Also check for 'choices' column
        if 'choices' in normalized_fieldnames:
            column_map['choices'] = normalized_fieldnames['choices']
        
        return column_map

    def process_excel_file(self, excel_file, processed_path, failed_path, sheet_name=None, header_row=1, start_row=2):
        """Process a single Excel file and import questions"""
        errors = []
        questions_created = 0
        questions_updated = 0
        choices_created = 0
        
        try:
            # Use pandas if available (better for reading Excel)
            if PANDAS_AVAILABLE:
                return self._process_with_pandas(excel_file, processed_path, failed_path, sheet_name, header_row, start_row)
            elif OPENPYXL_AVAILABLE:
                return self._process_with_openpyxl(excel_file, processed_path, failed_path, sheet_name, header_row, start_row)
            else:
                raise ValueError('No Excel library available')
                
        except Exception as e:
            error_msg = f'Error processing file: {str(e)}'
            self.stdout.write(self.style.ERROR(f'  {error_msg}'))
            
            # Move to failed folder
            error_log_path = failed_path / f'{excel_file.stem}_errors.txt'
            with open(error_log_path, 'w', encoding='utf-8', errors='replace') as log_file:
                safe_error_msg = str(error_msg).encode('utf-8', errors='replace').decode('utf-8')
                log_file.write(f'Error processing {excel_file.name}:\n\n{safe_error_msg}\n')
            
            dest_path = failed_path / excel_file.name
            if dest_path.exists():
                dest_path.unlink()
            excel_file.rename(dest_path)
            return False

    def _process_with_pandas(self, excel_file, processed_path, failed_path, sheet_name=None, header_row=1, start_row=2):
        """Process Excel file using pandas"""
        errors = []
        questions_created = 0
        questions_updated = 0
        choices_created = 0
        
        try:
            # Read Excel file
            excel_data = pd.read_excel(
                excel_file,
                sheet_name=sheet_name,
                header=header_row - 1,  # pandas uses 0-based indexing
                engine='openpyxl' if OPENPYXL_AVAILABLE else None
            )
            
            # If multiple sheets, get the first one or specified sheet
            if isinstance(excel_data, dict):
                if sheet_name:
                    df = excel_data.get(sheet_name)
                    if df is None:
                        raise ValueError(f'Sheet "{sheet_name}" not found in Excel file')
                else:
                    df = list(excel_data.values())[0]
            else:
                df = excel_data
            
            if df is None or df.empty:
                raise ValueError('Excel file is empty or no data found')
            
            # Get column names
            fieldnames = list(df.columns)
            
            # Map columns to standard format
            column_map = self.map_excel_columns(fieldnames)
            
            # Check if we have required columns
            if 'exam' not in column_map or 'question_text' not in column_map:
                raise ValueError(
                    f'Missing required columns. Found: {", ".join(fieldnames)}. '
                    f'Need: exam (or Exam) and question_text (or Question Text)'
                )
            
            # Check if we have any choice columns
            has_choice_columns = any(f'choice_{i}' in column_map for i in range(1, 7))
            has_choices_column = 'choices' in column_map
            
            if not (has_choice_columns or has_choices_column):
                raise ValueError(
                    'No choice columns found. Need either separate Choice 1, Choice 2, etc. columns '
                    'or a "choices" column with pipe-separated values'
                )
            
            # Process each row
            with transaction.atomic():
                for idx, row in df.iterrows():
                    row_num = start_row + idx  # Actual row number in Excel (1-based)
                    
                    # Skip empty rows
                    if row.isna().all():
                        continue
                    
                    try:
                        # Convert row to dictionary with mapped column names
                        mapped_row = {}
                        for standard_name, original_name in column_map.items():
                            value = row.get(original_name, '')
                            
                            # Convert pandas types to string/clean values
                            if pd.isna(value):
                                value = ''
                            elif isinstance(value, (int, float)):
                                # Keep numbers as strings to preserve formatting
                                value = str(value)
                            else:
                                value = str(value).strip()
                            
                            mapped_row[standard_name] = value
                        
                        # Remove brackets from text fields (if any)
                        text_fields = ['question_text', 'explanation', 'category', 'exam_category', 'exam_name', 'exam', 'topic']
                        for field in text_fields:
                            if field in mapped_row and isinstance(mapped_row[field], str):
                                mapped_row[field] = mapped_row[field].replace('[', '').replace(']', '')
                        
                        # Remove brackets from all choice fields
                        for i in range(1, 7):
                            choice_key = f'choice_{i}'
                            if choice_key in mapped_row and isinstance(mapped_row[choice_key], str):
                                mapped_row[choice_key] = mapped_row[choice_key].replace('[', '').replace(']', '')
                        
                        # Remove brackets from correct_answer
                        if 'correct_answer' in mapped_row and isinstance(mapped_row['correct_answer'], str):
                            mapped_row['correct_answer'] = mapped_row['correct_answer'].replace('[', '').replace(']', '')
                        
                        result = self.process_row(mapped_row, row_num, has_choice_columns)
                        if result['created']:
                            questions_created += 1
                        else:
                            questions_updated += 1
                        choices_created += result['choices_count']
                    except Exception as e:
                        error_msg = f'Row {row_num}: {str(e)}'
                        errors.append(error_msg)
                        try:
                            self.stdout.write(self.style.ERROR(f'  {error_msg}'))
                        except UnicodeEncodeError:
                            safe_msg = error_msg.encode('ascii', 'ignore').decode('ascii')
                            self.stdout.write(self.style.ERROR(f'  {safe_msg}'))
            
            # Process results
            if errors:
                # Move to failed folder with error log
                error_log_path = failed_path / f'{excel_file.stem}_errors.txt'
                with open(error_log_path, 'w', encoding='utf-8', errors='replace') as log_file:
                    log_file.write(f'Errors processing {excel_file.name}:\n\n')
                    for error in errors:
                        safe_error = str(error).encode('utf-8', errors='replace').decode('utf-8')
                        log_file.write(f'{safe_error}\n')
                
                dest_path = failed_path / excel_file.name
                if dest_path.exists():
                    dest_path.unlink()
                excel_file.rename(dest_path)
                self.stdout.write(self.style.WARNING(
                    f'  Moved to failed folder. {questions_created + questions_updated} questions processed, '
                    f'{len(errors)} errors encountered.'
                ))
                return False
            else:
                # Move to processed folder
                dest_path = processed_path / excel_file.name
                if dest_path.exists():
                    dest_path.unlink()
                excel_file.rename(dest_path)
                self.stdout.write(self.style.SUCCESS(
                    f'  Created: {questions_created}, Updated: {questions_updated}, '
                    f'Choices: {choices_created}'
                ))
                return True
                
        except Exception as e:
            raise

    def _process_with_openpyxl(self, excel_file, processed_path, failed_path, sheet_name=None, header_row=1, start_row=2):
        """Process Excel file using openpyxl"""
        errors = []
        questions_created = 0
        questions_updated = 0
        choices_created = 0
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            
            # Get sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f'Sheet "{sheet_name}" not found in Excel file. Available sheets: {wb.sheetnames}')
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Get header row
            header_row_data = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
            fieldnames = [str(cell) if cell is not None else '' for cell in header_row_data]
            
            # Map columns to standard format
            column_map = self.map_excel_columns(fieldnames)
            
            # Check if we have required columns
            if 'exam' not in column_map or 'question_text' not in column_map:
                raise ValueError(
                    f'Missing required columns. Found: {", ".join(fieldnames)}. '
                    f'Need: exam (or Exam) and question_text (or Question Text)'
                )
            
            # Check if we have any choice columns
            has_choice_columns = any(f'choice_{i}' in column_map for i in range(1, 7))
            has_choices_column = 'choices' in column_map
            
            if not (has_choice_columns or has_choices_column):
                raise ValueError(
                    'No choice columns found. Need either separate Choice 1, Choice 2, etc. columns '
                    'or a "choices" column with pipe-separated values'
                )
            
            # Process each row
            with transaction.atomic():
                for row_idx, row_data in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                    # Skip empty rows
                    if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row_data):
                        continue
                    
                    try:
                        # Create row dictionary
                        row_dict = {}
                        for col_idx, cell_value in enumerate(row_data):
                            if col_idx < len(fieldnames):
                                row_dict[fieldnames[col_idx]] = str(cell_value) if cell_value is not None else ''
                        
                        # Convert row to use mapped column names
                        mapped_row = {}
                        for standard_name, original_name in column_map.items():
                            value = row_dict.get(original_name, '')
                            
                            # Clean value
                            if isinstance(value, str):
                                value = value.strip()
                            
                            mapped_row[standard_name] = value
                        
                        # Remove brackets from text fields
                        text_fields = ['question_text', 'explanation', 'category', 'exam', 'topic']
                        for field in text_fields:
                            if field in mapped_row and isinstance(mapped_row[field], str):
                                mapped_row[field] = mapped_row[field].replace('[', '').replace(']', '')
                        
                        # Remove brackets from all choice fields
                        for i in range(1, 7):
                            choice_key = f'choice_{i}'
                            if choice_key in mapped_row and isinstance(mapped_row[choice_key], str):
                                mapped_row[choice_key] = mapped_row[choice_key].replace('[', '').replace(']', '')
                        
                        # Remove brackets from correct_answer
                        if 'correct_answer' in mapped_row and isinstance(mapped_row['correct_answer'], str):
                            mapped_row['correct_answer'] = mapped_row['correct_answer'].replace('[', '').replace(']', '')
                        
                        result = self.process_row(mapped_row, row_idx, has_choice_columns)
                        if result['created']:
                            questions_created += 1
                        else:
                            questions_updated += 1
                        choices_created += result['choices_count']
                    except Exception as e:
                        error_msg = f'Row {row_idx}: {str(e)}'
                        errors.append(error_msg)
                        try:
                            self.stdout.write(self.style.ERROR(f'  {error_msg}'))
                        except UnicodeEncodeError:
                            safe_msg = error_msg.encode('ascii', 'ignore').decode('ascii')
                            self.stdout.write(self.style.ERROR(f'  {safe_msg}'))
            
            wb.close()
            
            # Process results
            if errors:
                # Move to failed folder with error log
                error_log_path = failed_path / f'{excel_file.stem}_errors.txt'
                with open(error_log_path, 'w', encoding='utf-8', errors='replace') as log_file:
                    log_file.write(f'Errors processing {excel_file.name}:\n\n')
                    for error in errors:
                        safe_error = str(error).encode('utf-8', errors='replace').decode('utf-8')
                        log_file.write(f'{safe_error}\n')
                
                dest_path = failed_path / excel_file.name
                if dest_path.exists():
                    dest_path.unlink()
                excel_file.rename(dest_path)
                self.stdout.write(self.style.WARNING(
                    f'  Moved to failed folder. {questions_created + questions_updated} questions processed, '
                    f'{len(errors)} errors encountered.'
                ))
                return False
            else:
                # Move to processed folder
                dest_path = processed_path / excel_file.name
                if dest_path.exists():
                    dest_path.unlink()
                excel_file.rename(dest_path)
                self.stdout.write(self.style.SUCCESS(
                    f'  Created: {questions_created}, Updated: {questions_updated}, '
                    f'Choices: {choices_created}'
                ))
                return True
                
        except Exception as e:
            raise

    def remove_all_brackets(self, text):
        """Remove all square brackets from text"""
        if not isinstance(text, str):
            return text
        return text.replace('[', '').replace(']', '')

    def process_row(self, row, row_num, has_choice_columns=False):
        """Process a single row and create/update question - same logic as CSV import"""
        # Get or create category
        exam_category = self.remove_all_brackets(row.get('exam_category', 'AAAA').strip() or 'AAAA')
        category_name = row.get('category', 'General').strip()
        if not category_name:
            category_name = 'General'
        category_name = self.remove_all_brackets(category_name)
        category, _ = Category.objects.get_or_create(
            name=category_name,
            defaults={
                'exam_category': exam_category,
                'description': f'Category for {category_name} exams'
            }
        )
        # Update exam_category if it was provided and different
        if exam_category != 'AAAA' and category.exam_category == 'AAAA':
            category.exam_category = exam_category
            category.save()
        
        # Get or create exam
        exam_name_field = self.remove_all_brackets(row.get('exam_name', 'UPSC').strip() or 'UPSC')
        exam_set_name = row.get('exam', '').strip()
        if not exam_set_name:
            raise ValueError('exam name (set name) cannot be empty')
        exam_set_name = self.remove_all_brackets(exam_set_name)
        
        exam, _ = Exam.objects.get_or_create(
            category=category,
            name=exam_set_name,
            defaults={
                'exam_name': exam_name_field,
                'description': self.remove_all_brackets(row.get('exam_description', '')),
                'duration_minutes': int(row.get('duration_minutes', 60)) if row.get('duration_minutes', '').strip() else 60,
                'total_questions': int(row.get('total_questions', 10)) if row.get('total_questions', '').strip() else 10,
                'passing_score': int(row.get('passing_score', 60)) if row.get('passing_score', '').strip() else 60,
            }
        )
        # Update exam_name if it was provided and different
        if exam_name_field != 'UPSC' and exam.exam_name == 'UPSC':
            exam.exam_name = exam_name_field
            exam.save()
        
        # Get or create topic (optional)
        topic = None
        if row.get('topic', '').strip():
            topic_name = self.remove_all_brackets(row['topic'].strip())
            topic, _ = Topic.objects.get_or_create(
                exam=exam,
                name=topic_name,
                defaults={
                    'description': self.remove_all_brackets(row.get('topic_description', '')),
                    'order': int(row.get('topic_order', 0)) if row.get('topic_order', '').strip() else 0
                }
            )
        
        # Parse question data
        question_text = self.remove_all_brackets(row.get('question_text', '').strip())
        if not question_text:
            raise ValueError('question_text cannot be empty')
        
        question_type = row.get('question_type', 'single').strip().lower()
        if question_type not in ['single', 'multiple', 'true_false']:
            question_type = 'single'
        
        difficulty = row.get('difficulty', 'medium').strip().lower()
        if difficulty not in ['easy', 'medium', 'hard']:
            difficulty = 'medium'
        
        explanation = self.remove_all_brackets(row.get('explanation', '').strip())
        
        # Safely parse points
        points_str = str(row.get('points', '1')).strip()
        try:
            points = int(float(points_str)) if points_str else 1  # Handle float strings
        except (ValueError, TypeError):
            points = 1
        
        # Safely parse order
        order_str = str(row.get('order', '0')).strip()
        try:
            order = int(float(order_str)) if order_str else 0
        except (ValueError, TypeError):
            order = 0
        
        is_active = str(row.get('is_active', 'true')).strip().lower() in ['true', '1', 'yes', 'y', '']
        
        # Create or update question
        question, created = Question.objects.update_or_create(
            exam=exam,
            question_text=question_text,
            defaults={
                'topic': topic,
                'question_type': question_type,
                'difficulty': difficulty,
                'explanation': explanation,
                'points': points,
                'order': order,
                'is_active': is_active,
            }
        )
        
        # Parse choices
        choices_list = []
        
        if has_choice_columns:
            # Format with separate columns: Choice 1, Choice 2, etc.
            for i in range(1, 7):
                choice_key = f'choice_{i}'
                if choice_key in row and str(row[choice_key]).strip():
                    choice_text = self.remove_all_brackets(str(row[choice_key]).strip())
                    if choice_text:
                        choices_list.append(choice_text)
        else:
            # Format with pipe-separated choices column
            choices_text = str(row.get('choices', '')).strip()
            if choices_text:
                choices_text = self.remove_all_brackets(choices_text)
                choices_list = [self.remove_all_brackets(c.strip()) for c in choices_text.split('|') if c.strip()]
        
        if not choices_list:
            raise ValueError('At least one choice is required')
        
        # Parse correct answer(s) - same logic as CSV import
        correct_answer_text = self.remove_all_brackets(
            str(row.get('correct_answer', '')).strip() or str(row.get('correct_choices', '')).strip()
        )
        correct_indices = []
        
        if correct_answer_text:
            # Try to match by exact text first (case-insensitive, trimmed)
            for idx, choice_text in enumerate(choices_list):
                if choice_text.strip().lower() == correct_answer_text.strip().lower():
                    correct_indices.append(idx)
                    break
            
            # If not found by exact text, try partial/fuzzy matching
            if not correct_indices:
                correct_lower = correct_answer_text.strip().lower()
                for idx, choice_text in enumerate(choices_list):
                    choice_lower = choice_text.strip().lower()
                    if (correct_lower in choice_lower or choice_lower in correct_lower):
                        correct_indices.append(idx)
                        if question_type == 'single':
                            break
            
            # If still not found, try parsing as 1-indexed choice number
            if not correct_indices:
                try:
                    choice_num = int(float(correct_answer_text.strip()))
                    if 1 <= choice_num <= len(choices_list):
                        correct_indices.append(choice_num - 1)
                except ValueError:
                    # Try comma-separated indices
                    try:
                        indices = [int(float(x.strip())) for x in correct_answer_text.split(',')]
                        correct_indices = [idx - 1 for idx in indices if 1 <= idx <= len(choices_list)]
                    except ValueError:
                        pass
        
        if not correct_indices:
            error_msg = f'Could not determine correct answer. Provided: "{correct_answer_text}". Choices: {choices_list}'
            raise ValueError(error_msg)
        
        # Validate indices
        for idx in correct_indices:
            if idx < 0 or idx >= len(choices_list):
                raise ValueError(f'Invalid correct choice index: {idx + 1}')
        
        # Delete existing choices and create new ones
        question.choices.all().delete()
        choices_count = 0
        
        for idx, choice_text in enumerate(choices_list):
            is_correct = idx in correct_indices
            Choice.objects.create(
                question=question,
                choice_text=choice_text,
                is_correct=is_correct,
                order=idx
            )
            choices_count += 1
        
        return {
            'created': created,
            'choices_count': choices_count
        }


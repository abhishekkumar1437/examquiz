"""
Django management command to create a sample Excel file for question import
"""
from pathlib import Path
from django.core.management.base import BaseCommand
from django.conf import settings

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class Command(BaseCommand):
    help = 'Create a sample Excel file in the inbox folder for question import'

    def handle(self, *args, **options):
        if not PANDAS_AVAILABLE and not OPENPYXL_AVAILABLE:
            self.stdout.write(self.style.ERROR(
                'Error: Neither pandas nor openpyxl is installed. '
                'Please install one: pip install pandas openpyxl'
            ))
            return

        # Get paths
        base_dir = Path(settings.BASE_DIR)
        inbox_path = base_dir / 'inbox'
        inbox_path.mkdir(exist_ok=True)
        
        sample_file = inbox_path / 'sample_questions.xlsx'
        
        # Sample data
        sample_data = {
            'Category': [
                'Mathematics', 'Science', 'History', 'Geography', 'Science',
                'Mathematics', 'Science', 'History'
            ],
            'Exam': [
                'Basic Math Test', 'Chemistry Quiz', 'World History', 'World Geography',
                'Physics Quiz', 'Advanced Math', 'Biology Quiz', 'Ancient History'
            ],
            'Question Text': [
                'What is 2 + 2?',
                'What is the chemical formula for water?',
                'Who was the first President of India?',
                'Which is the largest ocean?',
                'What is the speed of light?',
                'Which numbers are prime?',
                'The Earth is round',
                'In which year did World War II end?'
            ],
            'Choice 1': [
                '3', 'H2O', 'Mahatma Gandhi', 'Atlantic', '300,000 km/s',
                '2', 'True', '1943'
            ],
            'Choice 2': [
                '4', 'CO2', 'Jawaharlal Nehru', 'Pacific', '150,000 km/s',
                '4', 'False', '1944'
            ],
            'Choice 3': [
                '5', 'O2', 'Dr. Rajendra Prasad', 'Indian', '500,000 km/s',
                '7', '', '1945'
            ],
            'Choice 4': [
                '6', 'NaCl', 'Sardar Patel', 'Arctic', '1,000,000 km/s',
                '9', '', '1946'
            ],
            'Correct Answer': [
                '4', 'H2O', 'Dr. Rajendra Prasad', 'Pacific', '300,000 km/s',
                '1,3', 'True', '1945'
            ],
            'Explanation': [
                'Simple addition: 2 + 2 = 4',
                'Water is H2O - two hydrogen atoms and one oxygen atom',
                'Dr. Rajendra Prasad was India\'s first President',
                'The Pacific Ocean is the largest ocean on Earth',
                'The speed of light in vacuum is approximately 300,000 km/s',
                '2 and 7 are prime numbers, 4 and 9 are composite',
                'The Earth is approximately spherical in shape',
                'World War II ended in 1945'
            ],
            'Difficulty': [
                'easy', 'easy', 'medium', 'easy', 'medium',
                'medium', 'easy', 'medium'
            ],
            'Topic': [
                'Arithmetic', 'Chemistry', 'Indian History', 'Geography',
                'Physics', 'Number Theory', 'Geography', 'World History'
            ],
            'Points': [1, 1, 1, 1, 1, 2, 1, 1],
            'Question Type': [
                'single', 'single', 'single', 'single', 'single',
                'multiple', 'true_false', 'single'
            ]
        }
        
        # Create Excel file using pandas if available, otherwise openpyxl
        if PANDAS_AVAILABLE:
            df = pd.DataFrame(sample_data)
            df.to_excel(sample_file, index=False, engine='openpyxl')
            self.stdout.write(self.style.SUCCESS(f'✅ Sample Excel file created successfully using pandas!'))
        elif OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Questions"
            
            # Write headers with styling
            headers = list(sample_data.keys())
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data
            num_rows = len(sample_data[headers[0]])
            for row_num in range(num_rows):
                for col_num, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row_num + 2, column=col_num)
                    cell.value = sample_data[header][row_num]
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Adjust column widths
            column_widths = {'A': 15, 'B': 18, 'C': 40, 'D': 20, 'E': 20, 'F': 20, 'G': 20,
                           'H': 18, 'I': 50, 'J': 12, 'K': 18, 'L': 8, 'M': 15}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            ws.freeze_panes = 'A2'
            wb.save(sample_file)
            self.stdout.write(self.style.SUCCESS(f'✅ Sample Excel file created successfully using openpyxl!'))
        
        self.stdout.write(self.style.SUCCESS(f'📁 Location: {sample_file}'))
        self.stdout.write(self.style.SUCCESS(f'📊 Contains {len(sample_data["Category"])} sample questions'))
        self.stdout.write(self.style.SUCCESS(f'\nColumns included:'))
        for col in sample_data.keys():
            self.stdout.write(f'  - {col}')
        self.stdout.write(self.style.SUCCESS(f'\nTo import this file, run:'))
        self.stdout.write(self.style.SUCCESS(f'  python manage.py import_questions_excel'))

